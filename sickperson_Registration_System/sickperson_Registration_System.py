from tkinter import *
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl,xlrd
from openpyxl import Workbook
import pathlib

background ="#BBDEFB"
framebg="#ECEFF1" #พื้นหลังในกรอบ
framefg="#212121" #ตัวอักษรในกรอบ


root =Tk()
root.title("ระบบบันทึกข้อมูลผู้ป่วย")
root.geometry("1250x700+210+100")
root.config(bg=background)
file=pathlib.Path('sickperson_data.xlsx')
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']="Registration No."
    sheet['B1']="เลขบัตรประจำตัวประชาชน"
    sheet['C1']="ชื่อ-นามสกุล"
    sheet['D1']="หมู่เลือด"
    sheet['E1']="เพศ"
    sheet['F1']="วัน/เดือน/ปีเกิด"
    sheet['G1']="วันที่สมัคร"
    sheet['H1']="สัญชาติ"
    sheet['I1']="อายุ"
    sheet['J1']="บ้านเลขที่/หมู่]"
    sheet['K1']="ตำบล"
    sheet['L1']="อำเภอ"
    sheet['M1']="จังหวัด"
    sheet['N1']="เบอรโทร"
    sheet['O1']="โรคประจำตัว"
    sheet['P1']="แพ้ยา"
    sheet['Q1']="ประวัติการรักษา"
    file.save('sickperson_data.xlsx')

#Exit window
def Exit():
    root.destroy()
#showImage
def showImage():
    global filename
    global img
    filename=filedialog.askopenfilename(initialdir=os.getcwd(),
                                        title="Select image file",filetype=(("JPG File","*.jpg"),
                                                                             ("PNG File","*.png"),
                                                                             ("All Files","*.txt")))
    img =(Image.open(filename))
    resized_image=img.resize((190,190))
    photo2 = ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image=photo2
#Registration NO.
def registration_no():
    file=openpyxl.load_workbook('sickperson_data.xlsx')
    sheet=file.active
    row=sheet.max_row

    max_row_value=sheet.cell(row=row,column=1).value

    try:
        Registration.set(max_row_value+1)
    except:
        Registration.set("1")
###Clear
def Clear():
    global img
    Id_user.set('')
    Name.set('')
    DOB.set('')
    Nationality.set('')
    Age.set('')
    Address.set('')
    Tambon.set('')
    Amphoe.set('')
    Changwat.set('')
    Group.set("เลือกหมู่เลือด")
    Tel.set('')
    Congenital_disease.set('')
    Drug_allergy.set('')
    txt_Box.delete("1.0","end")
    registration_no()

    saveButton.config(state='normal')
    img1=PhotoImage(file='Images/person.png')
    lbl.config(image=img1)
    lbl.image=img1
    img=""
#save
def Save():
    Registration_=Registration.get()
    Id_ = Id_user.get()
    Name_=Name.get()
    Group_=Group.get()
    try:
        G1=gender
    except:
        messagebox.showerror("error","Select Gender!")
        
    D2=DOB.get()
    Date_=Date.get()
    Nationality_=Nationality.get()
    Age_=Age.get()
    Address_=Address.get()
    Tambon_=Tambon.get()
    Amphoe_=Amphoe.get()
    Changwat_=Changwat.get()
    Tel_ =Tel.get()
    Congenital_disease_ =Congenital_disease.get()
    Drug_allergy_ =Drug_allergy.get()
    History_ = txt_Box.get("1.0","end-1c")
    
    if Name_=="" or Group_=="เลือกหมู่เลือด" or Nationality_=="" or Age_=="" or Address_=="" or Tambon_==""or Amphoe_=="" or Changwat_=="":
        messagebox.showerror("error","Few Data is missing!")
    else:
        file=openpyxl.load_workbook('sickperson_data.xlsx')
        sheet=file.active
        sheet.cell(column=1,row=sheet.max_row+1,value=Registration_)
        sheet.cell(column=2,row=sheet.max_row,value=Id_)
        sheet.cell(column=3,row=sheet.max_row,value=Name_)
        sheet.cell(column=4,row=sheet.max_row,value=Group_)
        sheet.cell(column=5,row=sheet.max_row,value=G1)
        sheet.cell(column=6,row=sheet.max_row,value=D2)
        sheet.cell(column=7,row=sheet.max_row,value=Date_)
        sheet.cell(column=8,row=sheet.max_row,value=Nationality_)
        sheet.cell(column=9,row=sheet.max_row,value=Age_)
        sheet.cell(column=10,row=sheet.max_row,value=Address_)
        sheet.cell(column=11,row=sheet.max_row,value=Tambon_)
        sheet.cell(column=12,row=sheet.max_row,value=Amphoe_)
        sheet.cell(column=13,row=sheet.max_row,value=Changwat_)
        sheet.cell(column=14,row=sheet.max_row,value=Tel_)
        sheet.cell(column=15,row=sheet.max_row,value=Congenital_disease_)
        sheet.cell(column=16,row=sheet.max_row,value=Drug_allergy_)
        sheet.cell(column=17,row=sheet.max_row,value=History_)
        
        file.save(r'sickperson_data.xlsx')
        try:
            img.save("sickperson Images/"+str(Registration_)+".jpg")
        except:
            messagebox.showinfo("info","Profile Picture is not available!!!")
        messagebox.showinfo("info","Susessfully data entered!!")
        Clear()
        registration_no()
        
#gender
def selection():
    global gender
    value=radio.get()
    if value==1:
        gender="Male"
        #print(gender)
    else:
        gender="Female"
        #print(gender)
#Search
def search():
    text = Search.get()
    Clear()
    saveButton.config(state='disable')
    file=openpyxl.load_workbook("sickperson_data.xlsx")
    sheet=file.active
    for row in sheet.rows:
        if row[1].value == text:
            name=row[1]
            print(str(name))
            reg_name_position=str(name)[14:-1]
            reg_name=str(name)[15:-1]
            print(reg_name_position)
            print(reg_name)
    try:
        print(str(name))
        #reg_name_position show A2,A3,A4......An
        x1=sheet.cell(row=int(reg_name),column=1).value
        x2=sheet.cell(row=int(reg_name),column=2).value
        x3=sheet.cell(row=int(reg_name),column=3).value
        x4=sheet.cell(row=int(reg_name),column=4).value
        x5=sheet.cell(row=int(reg_name),column=5).value
        x6=sheet.cell(row=int(reg_name),column=6).value
        x7=sheet.cell(row=int(reg_name),column=7).value
        x8=sheet.cell(row=int(reg_name),column=8).value
        x9=sheet.cell(row=int(reg_name),column=9).value
        x10=sheet.cell(row=int(reg_name),column=10).value
        x11=sheet.cell(row=int(reg_name),column=11).value
        x12=sheet.cell(row=int(reg_name),column=12).value
        x13=sheet.cell(row=int(reg_name),column=13).value
        x14=sheet.cell(row=int(reg_name),column=14).value
        x15=sheet.cell(row=int(reg_name),column=15).value
        x16=sheet.cell(row=int(reg_name),column=16).value
        x17=sheet.cell(row=int(reg_name),column=17).value
        print(x1)
        print(x2)
        print(x3)
        print(x4)
        print(x5)
        print(x6)
        print(x7)
        print(x8)
        print(x9)
        print(x10)
        print(x11)
        print(x12)
        print(x13)
        print(x14)
        print(x15)
        print(x16)
        print(x17)
        Registration.set(x1)
        Id_user.set(x2)
        Name.set(x3)
        Group.set(x4)
        if x5=='Female':
            R2.select()
        else:
            R1.select()
        DOB.set(x6)
        Date.set(x7)
        Nationality.set(x8)
        Age.set(x9)
        Address.set(x10)
        Tambon.set(x11)
        Amphoe.set(x12)
        Changwat.set(x13)
        Tel.set(x14)
        Congenital_disease.set(x15)
        Drug_allergy.set(x16)
        txt_Box.insert("end",x17)
        try:
            img = (Image.open("sickperson Images/"+str(x1)+".jpg"))
            resized_image=img.resize((190,190))
            photo2 = ImageTk.PhotoImage(resized_image)
            lbl.config(image=photo2)
            lbl.image=photo2
        except:
            pass 
    except:
        messagebox.showerror("Invalid","Invalid registration number!!!")
    

#update
def update():
    Registration_=Registration.get()
    Id_ = Id_user.get()
    Name_=Name.get()
    Group_=Group.get()
    selection()
    print(gender)
    G1=gender
    D2=DOB.get()
    Date_=Date.get()
    Nationality_=Nationality.get()
    Age_=Age.get()
    Address_=Address.get()
    Tambon_=Tambon.get()
    Amphoe_=Amphoe.get()
    Changwat_=Changwat.get()
    Tel_ =Tel.get()
    Congenital_disease_ =Congenital_disease.get()
    Drug_allergy_ =Drug_allergy.get()
    History_ = txt_Box.get("1.0","end-1c")

    file=openpyxl.load_workbook('sickperson_data.xlsx')
    sheet=file.active

    for row in sheet.rows:
        if row[0].value ==Registration_:
            name=row[0]
            print(str(name))
            reg_name_position=str(name)[14:-1]
            reg_name=str(name)[15:-1]
            print(reg_name)
    #sheet.cell(column=1,row=sheet.max_row,value=Registration_)
    sheet.cell(column=2,row=sheet.max_row,value=Id_)
    sheet.cell(column=3,row=sheet.max_row,value=Name_)
    sheet.cell(column=4,row=sheet.max_row,value=Group_)
    sheet.cell(column=5,row=sheet.max_row,value=G1)
    sheet.cell(column=6,row=sheet.max_row,value=D2)
    sheet.cell(column=7,row=sheet.max_row,value=Date_)
    sheet.cell(column=8,row=sheet.max_row,value=Nationality_)
    sheet.cell(column=9,row=sheet.max_row,value=Age_)
    sheet.cell(column=10,row=sheet.max_row,value=Address_)
    sheet.cell(column=11,row=sheet.max_row,value=Tambon_)
    sheet.cell(column=12,row=sheet.max_row,value=Amphoe_)
    sheet.cell(column=13,row=sheet.max_row,value=Changwat_)
    sheet.cell(column=14,row=sheet.max_row,value=Tel_)
    sheet.cell(column=15,row=sheet.max_row,value=Congenital_disease_)
    sheet.cell(column=16,row=sheet.max_row,value=Drug_allergy_)
    sheet.cell(column=17,row=sheet.max_row,value=History_)
    file.save(r'sickperson_data.xlsx')
    try:
        img.save("sickperson Images/"+str(Registration_)+".jpg")
    except:
        pass
    messagebox.showinfo("Update","Update Sucessfully!!")
    Clear()
                        
        
# Add image
#label = Label(root, image=bg)
#label.place(x = 0,y = 0)

#top frames
#Label(root,text="Email: jack12688jack@gmail.com",width=10,height=3,bg="#f0687C",anchor='e').pack(side=TOP,fill=X)
Label(root,text="ระบบบันทึกข้อมูลผู้ป่วย",width=10,height=2,bg="#0D47A1",fg='#fff',font='CmPrasanmit 25 bold').pack(side=TOP,fill=X) ###

#search box to update
Search = StringVar()
Entry(root,textvariable=Search,width=15,bd=2,font="CmPrasanmit 25").place(x=800,y=90)

imageicon3=PhotoImage(file="Images/Search2.png")
Srch=Button(root,text="Search",compound=LEFT,image=imageicon3,width=123,bg='#68ddfa',font="CmPrasanmit 13 bold",command=search)
Srch.place(x=1090,y=88)

imageicon4=PhotoImage(file="Images/Refresh.png")
Update_button=Button(root,image=imageicon4,bg="#06283D",command=update,width=45,height=45)
Update_button.place(x=730,y=87)        

#Registration and Date
Label(root,text="Registration No:",font="CmPrasanmit 16",fg='#212121',bg=background).place(x=30,y=95)
Label(root,text="Date:",font="CmPrasanmit 16",fg='#212121',bg=background).place(x=500,y=95)

Registration =IntVar()
Date = StringVar()

reg_entry = Entry(root,textvariable=Registration,width=15,font="CmPrasanmit 10")
reg_entry.place(x=160,y=100)

registration_no()

today =date.today()
d1 = today.strftime("%d/%m/%Y")
date_entry = Entry(root,textvariable=Date,width=15,font="CmPrasanmit 10")
date_entry.place(x=550,y=100)

Date.set(d1)

#Student details
obj =LabelFrame(root,text="ข้อมูลผู้ป่วย",font="CmPrasanmit 20",bd=2,width=900,bg=framebg,fg=framefg,height=340,relief=GROOVE)
obj.place(x=30,y=150)
Label(obj,text="เลขที่บัตรปีะจำตัวประชาชน:",font="CmPrasanmit 14",bg=framebg,fg=framefg).place(x=30,y=10)
Label(obj,text="ชื่อ-นามสกุล:",font="CmPrasanmit 14",bg=framebg,fg=framefg).place(x=30,y=40)
Label(obj,text="วัน/เดือน/ปีเกิด:",font="CmPrasanmit 14",bg=framebg,fg=framefg).place(x=30,y=80)
Label(obj,text="เพศ:",font="CmPrasanmit 14",bg=framebg,fg=framefg).place(x=30,y=120)

Label(obj,text="หมู่เลือด:",font="CmPrasanmit 14",bg=framebg,fg=framefg).place(x=500,y=40)
Label(obj,text="อายุ:",font="CmPrasanmit 14",bg=framebg,fg=framefg).place(x=500,y=80)
Label(obj,text="สัญชาติ:",font="CmPrasanmit 14",bg=framebg,fg=framefg).place(x=500,y=120)

Id_user=StringVar()
id_user = Entry(obj,textvariable=Id_user,width=40,font="CmPrasanmit 14")
id_user.place(x=230,y=10)

Name=StringVar()
name_entry = Entry(obj,textvariable=Name,width=20,font="CmPrasanmit 14")
name_entry.place(x=160,y=40)

DOB=StringVar()
dob_entry = Entry(obj,textvariable=DOB,width=20,font="CmPrasanmit 14")
dob_entry.place(x=160,y=80)

radio=IntVar()
R1 = Radiobutton(obj,text="ชาย",variable=radio,value=1,bg=framebg,font="CmPrasanmit 14",fg=framefg,command=selection)
R1.place(x=150,y=120)

R2 = Radiobutton(obj,text="หญิง",variable=radio,value=2,bg=framebg,font="CmPrasanmit 14",fg=framefg,command=selection)
R2.place(x=200,y=120)

Nationality=StringVar()
nationality_entry = Entry(obj,textvariable=Nationality,width=20,font="CmPrasanmit 14")
nationality_entry.place(x=630,y=120)

Age=StringVar()
age_entry = Entry(obj,textvariable=Age,width=20,font="CmPrasanmit 14")
age_entry.place(x=630,y=80)

Group=Combobox(obj,values=['A','B','AB','O'],font="CmPrasanmit 14",width=17,state="r")
Group.place(x=630,y=40)
Group.set("เลือกหมู่เลือด")

Label(obj,text="ที่อยู่",font="CmPrasanmit 14",bg=framebg,fg=framefg).place(x=10,y=160)
Label(obj,text="บ้านเลขที่/หมู่:",font="CmPrasanmit 14",bg=framebg,fg=framefg).place(x=30,y=200)
Label(obj,text="อำเภอ:",font="CmPrasanmit 14",bg=framebg,fg=framefg).place(x=30,y=240)
Label(obj,text="เบอร์โทร:",font="CmPrasanmit 14",bg=framebg,fg=framefg).place(x=30,y=280)

Address=StringVar()
address_entry =Entry(obj,textvariable=Address,width=20,font="CmPrasanmit 14")
address_entry.place(x=160,y=200)
Amphoe=StringVar()
amphoe_entry =Entry(obj,textvariable=Amphoe,width=20,font="CmPrasanmit 14")
amphoe_entry.place(x=160,y=240)
Tel=StringVar()
tel_entry =Entry(obj,textvariable=Tel,width=20,font="CmPrasanmit 14")
tel_entry.place(x=160,y=280)
Label(obj,text="ตำบล:",font="CmPrasanmit 14",bg=framebg,fg=framefg).place(x=500,y=200)
Label(obj,text="จังหวัด:",font="CmPrasanmit 14",bg=framebg,fg=framefg).place(x=500,y=240)
Tambon=StringVar()
tambon_entry =Entry(obj,textvariable=Tambon,width=20,font="CmPrasanmit 14")
tambon_entry.place(x=630,y=200)
Changwat=StringVar()
changwat_entry =Entry(obj,textvariable=Changwat,width=20,font="CmPrasanmit 14")
changwat_entry.place(x=630,y=240)

#Parents details
obj2 =LabelFrame(root,text="ข้อมูลการรักษา",font="CmPrasanmit 20",bd=2,width=900,bg=framebg,fg=framefg,height=170,relief=GROOVE)
obj2.place(x=30,y=505)

Label(obj2,text="โรคประจำตัว:",font="CmPrasanmit 14",bg=framebg,fg=framefg).place(x=30,y=10)

Congenital_disease=StringVar()
congenital_disease_entry =Entry(obj2,textvariable=Congenital_disease,width=20,font="CmPrasanmit 14")
congenital_disease_entry.place(x=160,y=10)

Label(obj2,text="แพ้ยา:",font="CmPrasanmit 14",bg=framebg,fg=framefg).place(x=500,y=10)
Drug_allergy=StringVar()
drug_allergy_entry =Entry(obj2,textvariable=Drug_allergy,width=20,font="CmPrasanmit 14")
drug_allergy_entry.place(x=630,y=10)

Label(obj2,text="ประวัติการรักษา:",font="CmPrasanmit 14",bg=framebg,fg=framefg).place(x=30,y=40)
txt_Box=Text(obj2,height =20,width=80)
txt_Box.place(x=160,y=40)


#History=StringVar()
#History_entry = Entry(obj2,textvariable=History,width=87,font="arial 10",)
#History_entry.place(x=160,y=40)
#image
f=Frame(root,bd=3,bg="black",width=235,height=235,relief=GROOVE)
f.place(x=980,y=150)

img=PhotoImage(file="Images/person.png")
lbl=Label(f,bg="black",image=img)
lbl.place(x=0,y=0)

#button

Button(root,text="Upload",width=19,height=2,font="arial 12 bold",bg="lightblue",command=showImage).place(x=1000,y=410)

saveButton=Button(root,text="Save",width=19,height=2,font="arial 12 bold",command=Save,bg="lightgreen")
saveButton.place(x=1000,y=480)

Button(root,text="Reset",width=19,height=2,font="arial 12 bold",bg="lightpink",command=Clear).place(x=1000,y=550)

Button(root,text="Exit",width=19,height=2,font="arial 12 bold",bg="grey",command=Exit).place(x=1000,y=620)
root.mainloop()



